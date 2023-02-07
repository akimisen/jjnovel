from openpyxl import Workbook,load_workbook
from collections.abc import Iterable
from datetime import datetime

def to_xlsx(data,savepath,sheetname=None,fields=None):
  wb=Workbook()
  if sheetname is None:
    ws=wb.active
  elif sheetname in wb.sheetnames:
    ws=wb[sheetname]
  else:
    ws=wb.create_sheet(sheetname)

  # insert fields
  if fields:
    ws.append(fields)

  # iter and insert data
  if isinstance(data, Iterable) and len(data)>0:
    for r in data:
      ws.append(r)

  wb.save(savepath)

def edit_xlsx(filepath,sheetname,data):
  wb=load_workbook(filepath)
  ws=wb[sheetname]
  if isinstance(data, Iterable) and len(data)>0:
    for r in data:
      ws.append(r)
  wb.save(filepath)

def xx_from_genre(genre):
  if genre.count('-')==3:
    _xx= genre.split('-')[1]
    rules = {'言情':'BG','纯爱':'BL'}
    if _xx in rules:
      return rules.get(_xx)
    return None

if __name__=='__main__':
  # aid=2210977
  # a=Author(aid)
  # date=datetime.now().strftime('%Y%m%d')
  # to_xlsx([],'data/authors_from_newDayList.xlsx',sheetname=date,fields=['id', 'name', 'follower_count', 'score', 'novels'])
  pass